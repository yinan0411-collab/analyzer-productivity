from __future__ import annotations

from io import BytesIO
from pathlib import Path
import re

import pandas as pd
import streamlit as st
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from attendance_analyzer import (
    Thresholds,
    analyze_attendance,
    build_summary,
    prepare_data,
    validate_columns,
)


st.set_page_config(
    page_title="员工班次偏差检查",
    page_icon="⏱️",
    layout="wide",
)

st.title("员工排班与实际打卡偏差检查")
st.caption(
    "上传考勤导出表，按日期和考勤组识别迟到、早退、提前上班、晚下班、缺卡、休息日打卡等需复核记录。"
)

with st.expander("判断口径说明", expanded=False):
    st.markdown(
        """
        - 程序不直接采用原表中的“异常状态”，而是重新比较计划时间和实际打卡时间。
        - 结果表示“需要复核的记录”，不等同于已经认定员工违规。
        - 只有超过阈值的记录才会显示；刚好等于阈值不会被标记。
        - 休息日有打卡、整班无打卡、有打卡无排班可单独开启或关闭。
        """
    )

uploaded_files = st.file_uploader(
    "上传考勤表（支持同时上传多个 Excel 文件）",
    type=["xlsx", "xls"],
    accept_multiple_files=True,
)

with st.sidebar:
    st.header("分析设置")

    use_same_threshold = st.toggle("四类偏差使用同一阈值", value=True)
    if use_same_threshold:
        common_threshold = st.number_input(
            "统一阈值（分钟）",
            min_value=0,
            max_value=720,
            value=30,
            step=5,
        )
        thresholds = Thresholds(
            early_start=int(common_threshold),
            late_start=int(common_threshold),
            early_leave=int(common_threshold),
            late_leave=int(common_threshold),
        )
    else:
        thresholds = Thresholds(
            early_start=int(st.number_input(
                "允许提前上班（分钟）", 0, 720, 30, 5
            )),
            late_start=int(st.number_input(
                "允许迟到（分钟）", 0, 720, 30, 5
            )),
            early_leave=int(st.number_input(
                "允许提前下班（分钟）", 0, 720, 30, 5
            )),
            late_leave=int(st.number_input(
                "允许晚下班（分钟）", 0, 720, 30, 5
            )),
        )

    st.divider()
    flag_missing = st.checkbox("标记缺卡/整班无打卡", value=True)
    flag_rest = st.checkbox("标记休息日打卡", value=True)
    flag_no_schedule = st.checkbox("标记有打卡无排班", value=True)
    deduplicate = st.checkbox(
        "合并文件后自动去重",
        value=True,
        help="按用户编码 + 日期 + 考勤组保留最新记录。",
    )


def read_uploaded_files(files) -> pd.DataFrame:
    frames = []
    for uploaded in files:
        uploaded.seek(0)
        try:
            frame = pd.read_excel(uploaded, sheet_name=0)
        except Exception as exc:
            st.error(f"无法读取 {uploaded.name}：{exc}")
            continue
        frame["来源文件"] = uploaded.name
        frames.append(frame)

    if not frames:
        return pd.DataFrame()

    return pd.concat(frames, ignore_index=True, sort=False)


def safe_sheet_name(value: object, used_names: set[str]) -> str:
    base = str(value) if value is not None else "未知日期"
    base = re.sub(r'[\[\]:*?/\\]', "-", base)[:31] or "明细"
    candidate = base
    index = 2
    while candidate in used_names:
        suffix = f"-{index}"
        candidate = f"{base[:31-len(suffix)]}{suffix}"
        index += 1
    used_names.add(candidate)
    return candidate


def format_workbook(writer: pd.ExcelWriter) -> None:
    workbook = writer.book
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center")

    for sheet in workbook.worksheets:
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions

        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

        for column_cells in sheet.columns:
            letter = get_column_letter(column_cells[0].column)
            max_length = 0
            for cell in column_cells[:3000]:
                value = "" if cell.value is None else str(cell.value)
                max_length = max(max_length, len(value))
            sheet.column_dimensions[letter].width = min(max(max_length + 2, 10), 40)

        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)

        header_map = {cell.value: cell.column for cell in sheet[1]}
        for name in [
            "计划上班时间", "实际上班时间",
            "计划下班时间", "实际下班时间",
        ]:
            if name in header_map:
                col_idx = header_map[name]
                for cell in sheet.iter_cols(
                    min_col=col_idx, max_col=col_idx, min_row=2
                ):
                    for item in cell:
                        item.number_format = "yyyy-mm-dd hh:mm"


def make_excel(result: pd.DataFrame, summary: pd.DataFrame) -> bytes:
    output = BytesIO()
    used_names: set[str] = set()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        summary.to_excel(writer, index=False, sheet_name="异常汇总")
        used_names.add("异常汇总")

        for date_value in sorted(result["日期"].dropna().unique()):
            daily = result[result["日期"] == date_value].copy()
            sheet_name = safe_sheet_name(date_value, used_names)
            daily.to_excel(writer, index=False, sheet_name=sheet_name)

        if result["日期"].isna().any():
            unknown = result[result["日期"].isna()].copy()
            unknown.to_excel(
                writer,
                index=False,
                sheet_name=safe_sheet_name("未知日期", used_names),
            )

        format_workbook(writer)

    output.seek(0)
    return output.getvalue()


if not uploaded_files:
    st.info("请先上传考勤导出表。")
    st.stop()

raw_data = read_uploaded_files(uploaded_files)
if raw_data.empty:
    st.stop()

prepared = prepare_data(raw_data, deduplicate=deduplicate)
missing_columns = validate_columns(prepared)
if missing_columns:
    st.error(
        "文件缺少必要字段："
        + "、".join(missing_columns)
        + "。请确认上传的是考勤明细导出表。"
    )
    st.write("当前识别到的列：", list(prepared.columns))
    st.stop()

result = analyze_attendance(
    prepared,
    thresholds=thresholds,
    flag_missing_punches=flag_missing,
    flag_rest_day_punches=flag_rest,
    flag_punch_without_schedule=flag_no_schedule,
)

if result.empty:
    st.success("按照当前阈值和规则，没有发现需要复核的员工记录。")
    st.stop()

all_groups = sorted(
    group for group in result["考勤组"].dropna().astype(str).unique()
)
all_dates = sorted(date for date in result["日期"].dropna().unique())

with st.sidebar:
    st.divider()
    st.header("结果筛选")
    selected_groups = st.multiselect(
        "考勤组",
        options=all_groups,
        default=all_groups,
    )

    if all_dates:
        selected_date_range = st.date_input(
            "日期范围",
            value=(all_dates[0], all_dates[-1]),
            min_value=all_dates[0],
            max_value=all_dates[-1],
        )
    else:
        selected_date_range = None

filtered = result[result["考勤组"].isin(selected_groups)].copy()

if selected_date_range:
    if isinstance(selected_date_range, tuple) and len(selected_date_range) == 2:
        start_date, end_date = selected_date_range
    else:
        start_date = end_date = selected_date_range
    filtered = filtered[
        filtered["日期"].between(start_date, end_date, inclusive="both")
    ]

summary = build_summary(filtered)

metric1, metric2, metric3, metric4 = st.columns(4)
metric1.metric("异常人次", f"{len(filtered):,}")
metric2.metric("异常员工数", f"{filtered['用户编码'].nunique():,}")
metric3.metric("涉及日期", f"{filtered['日期'].nunique():,}")
metric4.metric("涉及考勤组", f"{filtered['考勤组'].nunique():,}")

tab_daily, tab_summary = st.tabs(["按日期查看", "异常汇总"])

with tab_daily:
    available_dates = sorted(filtered["日期"].dropna().unique())
    if not available_dates:
        st.warning("当前筛选条件下没有记录。")
    else:
        selected_date = st.selectbox(
            "选择日期",
            options=available_dates,
            format_func=lambda value: str(value),
        )
        daily_data = filtered[filtered["日期"] == selected_date]

        for group_name, group_data in daily_data.groupby("考勤组", dropna=False):
            title = f"{group_name or '未填写考勤组'}（{len(group_data)} 人次）"
            with st.expander(title, expanded=True):
                st.dataframe(
                    group_data,
                    use_container_width=True,
                    hide_index=True,
                    column_config={
                        "上班偏差(分钟)": st.column_config.NumberColumn(
                            format="%.1f"
                        ),
                        "下班偏差(分钟)": st.column_config.NumberColumn(
                            format="%.1f"
                        ),
                    },
                )

with tab_summary:
    st.subheader("按日期、考勤组汇总")
    st.dataframe(summary, use_container_width=True, hide_index=True)

    if not summary.empty:
        date_chart = (
            summary.groupby("日期", as_index=False)["异常人次"]
            .sum()
            .set_index("日期")
        )
        st.subheader("每日异常人次")
        st.bar_chart(date_chart)

        group_chart = (
            summary.groupby("考勤组", as_index=False)["异常人次"]
            .sum()
            .sort_values("异常人次", ascending=False)
            .set_index("考勤组")
        )
        st.subheader("各考勤组异常人次")
        st.bar_chart(group_chart)

download_col1, download_col2 = st.columns(2)

with download_col1:
    excel_bytes = make_excel(filtered, summary)
    st.download_button(
        "下载异常 Excel（按日期分工作表）",
        data=excel_bytes,
        file_name="员工班次偏差异常明细.xlsx",
        mime=(
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        ),
        use_container_width=True,
    )

with download_col2:
    csv_bytes = filtered.to_csv(index=False).encode("utf-8-sig")
    st.download_button(
        "下载全部异常 CSV",
        data=csv_bytes,
        file_name="员工班次偏差异常明细.csv",
        mime="text/csv",
        use_container_width=True,
    )

st.caption(
    "建议先用 30–60 分钟阈值识别明显班次错位，再结合请假、临时调班、批准加班记录进行人工确认。"
)
